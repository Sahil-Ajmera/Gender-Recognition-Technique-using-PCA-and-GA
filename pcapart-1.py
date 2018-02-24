#To get the eigenvectors and eigenvalues from the training data
#Make changes as mentioned
import numpy as np
import openpyxl
from matlab.engine import *

#To find mean of all the values
'''
excel_document=openpyxl.load_workbook('testdata-1M.xlsx')#Make change here
sheet = excel_document.get_sheet_by_name('Sheet1')

multiple_cells = sheet['A1':'AHP1']
Value= ' '
for row in multiple_cells:
    for cell in row:
        Value=Value+cell.column+' '
RetrievedColumnValues=Value.split(' ')
multiple_cells=sheet['A404':'AHP404']#Make change here
k=1
for row in multiple_cells:
    for cell in row:
        FValue=str(RetrievedColumnValues[k])
        Fvalue1=FValue+'1'
        Fvalue2=FValue+'399' #Make change here 
        cell.value='=AVERAGE('+Fvalue1+':'+Fvalue2+')'
        k=k+1
excel_document.save(filename='testdata-1M.xlsx')#Make change here
print('Done.Now manually paste ')
'''
#Manually paste values after this from original excel sheet to new1 excel
#sheet

#To find the difference between mean and the other values

'''
excel_document=openpyxl.load_workbook('new1.xlsx')#Make change here
sheet = excel_document.get_sheet_by_name('Sheet1')
Value=' '
multiple_cells=sheet['A404':'AHP404']
for row in multiple_cells:
    for cell in row:
        Value=Value+str(cell.value)+' '
RetrievedValues=Value.split(' ')
RetrievedValues.pop()
RetrievedValues.pop(0)
print(RetrievedValues)
k=0
Val=' '

multiple_cells=sheet["A1:AHP399"] #Make changes here
for row in multiple_cells:
    for cell in row:
        x=float(RetrievedValues[k])-cell.value
        Val=Val+str(x)+' '
        k=k+1
    FinalVal=Val.split(' ')
    row=cell.row   
    newrow=row+410 #Make change here
    newinitial='A'+str(newrow)
    newfinal='AHP'+str(newrow)
    multiple_cells=sheet[newinitial:newfinal]
    l=1
    for row in multiple_cells:
        for cell in row:
            cell.value=FinalVal[l]
            l=l+1
    k=0
    Val=' '
excel_document.save(filename='new1.xlsx')
'''
listofzeroes=[0]*30
maximum=listofzeroes
#To convert the 1X900 to 30X30
eng=start_matlab()
excel_document=openpyxl.load_workbook('new1.xlsx')#Make change here
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells=sheet['A411':'AHP809'] #Make changes here

for row in multiple_cells:
    k=0
    Value=' '
    for cell in row:
        Value=Value+str(cell.value)+' '
        k=k+1
        if k==30:
            Value=Value+';'+' '
            k=0
    ReshapedValues=Value.split(' ')
    ReshapedValues.pop()
    ReshapedValues.pop(0)
    for i in range(len(ReshapedValues) - 1, -1, -1):  
    # iterate over reversed indices's
        if ReshapedValues[i] == ';':
            del ReshapedValues[i]
    print(ReshapedValues)
    
    a=np.reshape(ReshapedValues,(30,30))
    print('\n')
    print('\n')
    print(a)
    print('\n')
    data=np.array(a).astype(np.float)
    b=np.cov(data)
    print(b)
    print('\n')
    print(b.shape)
    e=np.linalg.eigvals(b)
    print(e)
    print(e.shape)
    if np.all(np.asarray(e) > np.asarray(listofzeroes)):

        maximum=e
print('Finalllly!!!')
print(maximum)
print(maximum.shape)
print(a.shape)


multiple_cells=sheet['A411':'AHP809']
for row in multiple_cells:
    k=0
    Value=' '
    for cell in row:
        Value=Value+str(cell.value)+' '
        k=k+1
        if k==30:
            Value=Value+';'+' '
            k=0
    ReshapedValues=Value.split(' ')
    ReshapedValues.pop()
    ReshapedValues.pop(0)
    for i in range(len(ReshapedValues) - 1, -1, -1):  
    # iterate over reversed indices's
        if ReshapedValues[i] == ';':
            del ReshapedValues[i]
    print(ReshapedValues)

    a=np.reshape(ReshapedValues,(30,30))
    final=(np.array(a).astype(np.float))*(np.array(maximum).astype(np.float))
    print("Say Hi to new feature vectors")
    print(final)
    
    








