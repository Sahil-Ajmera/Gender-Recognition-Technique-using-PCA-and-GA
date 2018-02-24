from __future__ import division
import openpyxl
import random
import csv
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

#Saving genetic data
#1
'''
excel_document=openpyxl.load_workbook('geneticdata.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells = sheet['A2':'AHP400']
for row in multiple_cells:
    for cell in row:
        cell.value=random.randint(0,1)
excel_document.save(filename='geneticdata.xlsx')
print('Genetic Data for the images have been saved')
print('transfer genetic data to AccordingToFeatures4  sheet 2And AccordingToFeatures5 and eigen project to sheet3 and accordingtofeatures5 sheet3')

#2

#Multiplying the genetic data with the eigen projects to get the eigen features for SVM training
#Multiply two xlsx files not csv files becoz openpyxl does not support it.
#change name of sheet from finalEigenProjects to sheet1

excel_document=openpyxl.load_workbook('geneticdata.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells = sheet['A1':'AHP399']

excel_document=openpyxl.load_workbook('AccordingToFeatures3.xlsx')#InitialDif
sheet1 = excel_document.get_sheet_by_name('Sheet2')
sheet2 = excel_document.get_sheet_by_name('Sheet3')
sheet3=excel_document.get_sheet_by_name('Sheet4')

k=0
value1=' '
multiple_cells=sheet1['A1':'AHP399']#Original:AHP399
for row in multiple_cells:
    for cell in row:
        value1=value1+str(cell.value)+' '
        k=k+1

multiple_cellss=sheet2['A1':'AHP399']#Original:AHP399
k=0
value2=' '
for row in multiple_cellss:
    for cell in row:
        value2=value2+str(cell.value)+' '
        k=k+1
        
RetrievedValues1=value1.split(' ')
RetrievedValues1.pop()
RetrievedValues1.pop(0)
a = np.array(RetrievedValues1)
a[a == ''] = 0.0
ReshapedValues1 = a.astype(np.float) 

RetrievedValues2=value2.split(' ')
RetrievedValues2.pop()
RetrievedValues2.pop(0)
a = np.array(RetrievedValues2)
a[a == ''] = 0.0
ReshapedValues2 = a.astype(np.float) 

k=0
multiple_cells=sheet3['A1':'AHP399']#Original:AHP399
for row in multiple_cells:
    for cell in row:
        cell.value=ReshapedValues1[k]*ReshapedValues2[k]
        k=k+1
print('done!Now Train the SVM.Convert to csv file and svm training ')
excel_document.save(filename='AccordingToFeatures3.xlsx')
 
#3
#SVM Training done ! 
#Now move on to genetic algorithm

#First iteration fitness values <0.22 selected

'''
a=0.9
b=0.1
SG=900
k=0
AR=0.78
FG=0
excel_document=openpyxl.load_workbook('AccordingToFeatures4.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells = sheet['A1':'OI900']

l=0
#for iteration in range(1,5):
#Removing ones with low fitness function
count=0
fitness=[]
wipeout=[]
for rows in multiple_cells:
    for cell in rows:
        if type(cell.value)=='Nonetype':
            continue
        if cell.value==1:
            FG=FG+1
    fitness.append(a*(1-AR)+b*(FG/SG))
    print(a*(1-AR)+b*(FG/SG))
    FG=0
excel_document=openpyxl.load_workbook('AccordingToFeatures5.xlsx')#SameData as 4
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells = sheet['A1':'OI900']

for index, item in enumerate(fitness):
                        if item<0.22:
                            count=count+1  
                        elif item>0.22:
                            wipeout.append(index+1)
print('Number of chromosomes with higher than weak fitness value:')
print(count)

for values in wipeout:
                        part1='A'+str(values)
                        part2='OI'+str(values)
                        multiple_cells=sheet[part1:part2]
                        for rows in multiple_cells:
                                                for cell in rows:
                                                                        cell.value=None
excel_document.save(filename='AccordingToFeatures5.xlsx')

#4


#Crossover operation
#Selecting Parents
a=0.9
b=0.1
SG=900
k=0
AR=0.78
FG=0
fitty=0.0
indices=[]
excel_document=openpyxl.load_workbook('AccordingToFeatures5.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells = sheet['A1':'OI900']
count=0
fitness=[]
wipeout=[]
for rows in multiple_cells:
    count=count+1
    for cell in rows:
            if cell.value==1:
                FG=FG+1
    if FG!=0:
        indices.append(cell.row)
        fitness.append(a*(1-AR)+b*(FG/SG))       
    FG=0
k=0




excel_document=openpyxl.load_workbook('FittingFunctionValues.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells=sheet['A1':'A421']#Make change here
for values in indices:
    
    sheet['A'+str(values)]=fitness[k]
    k=k+1

excel_document.save(filename='FittingFunctionValues.xlsx')
    
#5
#Mark the parents in the FittingFunctionValues excel file
excel_document=openpyxl.load_workbook('FittingFunctionValues.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells = sheet['A1':'A900']

Crossoverparents=[]
flag=[0]*20
for rows in multiple_cells:
    for cell in rows:
        if cell.value < 0.218 and cell.value!=None:
            Crossoverparents.append(cell.row)
            cell.font = Font(color=colors.RED, italic=True)
excel_document.save(filename='FittingFunctionValues.xlsx')
print('Number of parents involved in crossover where 2 at a time undergo crossover:')
print(len(Crossoverparents))


excel_document=openpyxl.load_workbook('AccordingToFeatures5.xlsx')
k=0
sheet = excel_document.get_sheet_by_name('Sheet1')
l=0
index=0
count=902
while index!=20:
        values=' '
        value12=' '
        firstparent='A'+str(Crossoverparents[index])
        half1='HO'+str(Crossoverparents[index])
        multiple_cells1=sheet[firstparent:half1]
        for rows in multiple_cells1:
                for cell in rows:
                        value12=value12+str(cell.value)+' '
        ReshapedValues=value12.split(' ')
        ReshapedValues.pop()
        ReshapedValues.pop(0)
        firstparent='A'+str(count)
        first_half='HO'+str(count)        
        multiple_cells_Final=sheet[firstparent:first_half]
        k=0
        for rows in multiple_cells_Final:
                for cell in rows:
                        cell.value=int(ReshapedValues[k])
                        k=k+1
        k=0
        
        secondparent='OI'+str(Crossoverparents[index+1])
        half2='HP'+str(Crossoverparents[index+1])
        multiple_cells2=sheet[half2:secondparent]
        for rows in multiple_cells2:
                for cell in rows:
                        values=values+str(cell.value)+' '
        
        ReshapedValues=values.split(' ')
        ReshapedValues.pop()
        ReshapedValues.pop(0)
        
        secondparent='HP'+str(count)
        second_half='OI'+str(count)
        multiple_cells_Final=sheet[secondparent:second_half]
        k=0
        for rows in multiple_cells_Final:
                for cell in rows:
                        cell.value=int(ReshapedValues[k])
                        k=k+1
        k=0
        count=count+1
        index=index+2
excel_document.save(filename='AccordingToFeatures5.xlsx')



#Mutation Operation
excel_document=openpyxl.load_workbook('AccordingToFeatures5.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
k=0
count2=902
index=0
count=0
while index!=20:
        startingindex='A'+str(count2)
        endingindex='OI'+ str(count2)
        multiple_cells=sheet[startingindex:endingindex]
        for rows in multiple_cells:
                for cell in rows:
                        if cell.value==1 and count==0:
                                
                                cell.value=0
                                count=1
        index=index+2
        count=0
        count2=count2+1
excel_document.save(filename='AccordingToFeatures5.xlsx')
#Now delete the unwanted rows in genetic data and eigen projects
#multiply both and train through the back propogation neural network

'''    
#6
parameters=[]
excel_document=openpyxl.load_workbook('AccordingToFeatures5.xlsx')#FinalEigenFeatureCalculation
sheet = excel_document.get_sheet_by_name('Sheet2')
sheet2 = excel_document.get_sheet_by_name('Sheet3')
multiple_cells=sheet['A1':'AHP399']
count=0
for rows in multiple_cells:
    for cell in rows:
        if cell.value==None:
            parameters.append(cell.column)          
    break        
print(parameters)

for letter in parameters:
    start=letter+str(1)
    end=letter+str(399)
    multiple_cells=sheet2[start:end]
    for rows in multiple_cells:
        for cell in rows:
            cell.value=None
excel_document.save(filename='AccordingToFeatures5.xlsx')#FinalEigenFeatureCalculation 

#Multiply sheet 2 and 3 again and train and test to back propagation neural network
#7
excel_document=openpyxl.load_workbook('AccordingToFeatures5.xlsx')#InitialDif
sheet1 = excel_document.get_sheet_by_name('Sheet2')
sheet2 = excel_document.get_sheet_by_name('Sheet3')
sheet3=excel_document.get_sheet_by_name('Sheet4')

k=0
value1=' '
multiple_cells=sheet1['A1':'PO399']#Original:AHP399
for row in multiple_cells:
    for cell in row:
        value1=value1+str(cell.value)+' '
        k=k+1

multiple_cellss=sheet2['A1':'PO399']#Original:AHP399
k=0
value2=' '
for row in multiple_cellss:
    for cell in row:
        value2=value2+str(cell.value)+' '
        k=k+1
        
RetrievedValues1=value1.split(' ')
RetrievedValues1.pop()
RetrievedValues1.pop(0)
a = np.array(RetrievedValues1)
a[a == ''] = 0.0
ReshapedValues1 = a.astype(np.float) 

RetrievedValues2=value2.split(' ')
RetrievedValues2.pop()
RetrievedValues2.pop(0)
a = np.array(RetrievedValues2)
a[a == ''] = 0.0
ReshapedValues2 = a.astype(np.float) 

k=0
multiple_cells=sheet3['A1':'PO399']#Original:AHP399
for row in multiple_cells:
    for cell in row:
        cell.value=ReshapedValues1[k]*ReshapedValues2[k]
        k=k+1
print('done!Now convert to csv file and train the back-propagation neural net ')
excel_document.save(filename='AccordingToFeatures5.xlsx')
'''           
        
            
