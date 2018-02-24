#To get the eigenvectors and eigenvalues from the training data
#Make changes as mentioned
import numpy as np
import openpyxl
from matlab.engine import *
import csv

#To find mean of all the values
#1
'''
excel_document=openpyxl.load_workbook('testdata-1M.xlsx')#Make change here
sheet = excel_document.get_sheet_by_name('Sheet1')

multiple_cells = sheet['A1':'AHP1']
Value= ' '
for row in multiple_cells:
    for cell in row:
        Value=Value+cell.column+' '
RetrievedColumnValues=Value.split(' ')
multiple_cells=sheet['A404':'AHP404']#Make change here
k=1
for row in multiple_cells:
    for cell in row:
        FValue=str(RetrievedColumnValues[k])
        Fvalue1=FValue+'1'
        Fvalue2=FValue+'399' #Make change here 
        cell.value='=AVERAGE('+Fvalue1+':'+Fvalue2+')'
        k=k+1
excel_document.save(filename='testdata-1M.xlsx')#Make change here
print('Done.Now manually paste ')

#Manually paste values after this from original excel sheet to new1 excel
#sheet

#To find the difference between mean and the other values
#2

excel_document=openpyxl.load_workbook('new1.xlsx')#Make change here
sheet = excel_document.get_sheet_by_name('Sheet1')
Value=' '
multiple_cells=sheet['A404':'AHP404']
for row in multiple_cells:
    for cell in row:
        Value=Value+str(cell.value)+' '
RetrievedValues=Value.split(' ')
RetrievedValues.pop()
RetrievedValues.pop(0)
k=0
Val=' '

multiple_cells=sheet["A1:AHP399"] #Make changes here
for row in multiple_cells:
    for cell in row:
        x=float(RetrievedValues[k])-cell.value
        Val=Val+str(x)+' '
        k=k+1
    FinalVal=Val.split(' ')
    row=cell.row   
    newrow=row+410 #Make change here
    newinitial='A'+str(newrow)
    newfinal='AHP'+str(newrow)
    multiple_cells=sheet[newinitial:newfinal]
    l=1
    for row in multiple_cells:
        for cell in row:
            cell.value=FinalVal[l]
            l=l+1
    k=0
    Val=' '
excel_document.save(filename='new1.xlsx')

'''
#3
f1=open('EigenProjects.csv','ab')
writer=csv.writer(f1)

#listofzeroes=[0]*30
#maximum=listofzeroes
maximum=np.zeros((30L,))

listofzeroes=np.zeros((30L,))

#To convert the 1X900 to 30X30
eng=start_matlab()
excel_document=openpyxl.load_workbook('new1.xlsx')#Make change here
sheet = excel_document.get_sheet_by_name('Sheet1')
multiple_cells=sheet['A411':'AHP809'] #Make changes here

for row in multiple_cells:
    k=0
    Value=' '
    for cell in row:
        Value=Value+str(cell.value)+' '
        k=k+1
        if k==30:
            Value=Value+';'+' '
            k=0
    ReshapedValues=Value.split(' ')
    ReshapedValues.pop()
    ReshapedValues.pop(0)
    for i in range(len(ReshapedValues) - 1, -1, -1):  
    # iterate over reversed indices's
        if ReshapedValues[i] == ';':
            del ReshapedValues[i]
    #print(ReshapedValues)
    #x = np.array(ReshapedValues, dtype='|S4')
    #ReshapedValues = x.astype(np.float)
    a = np.array(ReshapedValues)
    a[a == ''] = 0.0
    ReshapedValues = a.astype(np.float)  
    
    #a=np.reshape(ReshapedValues,(30,30))
    a=np.array(ReshapedValues)
    a.resize((30,30))
    #print(a.shape)
    #data=np.array(a).astype(np.float)
    data=np.matrix(a).astype(np.float)
    b=np.cov(data)
    e,w=np.linalg.eig(b)
    #print(e.shape)
    #b=np.array(e)
    #b.resize((30,1))
    #print(b.shape)
    #if np.all(np.asarray(e) > np.asarray(listofzeroes)):
    if np.all(np.matrix(e) > np.matrix(maximum)):
        maximum=e
        vector=w
print('Finalllly!!!')
print('Eigenvalue:')
print(maximum.shape)
print(maximum)
print('Eigenvector:')
print(vector.shape)
print(vector)
#print(w.shape)
#c=np.array(maximum)
#c.resize((30,1))
#print(c.shape)
count=0
multiple_cells=sheet['A411':'AHP809'] #Make changes here

for row in multiple_cells:
    k=0
    Value=' '
    for cell in row:
        Value=Value+str(cell.value)+' '
        k=k+1
        if k==30:
            Value=Value+';'+' '
            k=0
    count=count+1
    print(count)
    ReshapedValues=Value.split(' ')
    ReshapedValues.pop()
    ReshapedValues.pop(0)
    for i in range(len(ReshapedValues) - 1, -1, -1):  
    # iterate over reversed indices's
        if ReshapedValues[i] == ';':
            del ReshapedValues[i]
    #print(ReshapedValues)
    #x = np.array(ReshapedValues, dtype='|S4')
    #ReshapedValues = x.astype(np.float)
    a = np.array(ReshapedValues)
    a[a == ''] = 0.0
    ReshapedValues = a.astype(np.float)    
    #a=np.reshape(ReshapedValues,(30,30))
    a=np.array(ReshapedValues)
    a.resize((30,30))
    #print(a.shape)
    #final=(np.array(a).astype(np.float))*(np.array(maximum).astype(np.float))
    final=np.dot(a,vector)
    #print(final.shape)
    #print("Say Hi to new feature vectors")
    print(final)

    b=np.array(final)
    b.resize((1,900))
    '''
    for values in b:
        writer.writerow(values)
f1.close()
'''